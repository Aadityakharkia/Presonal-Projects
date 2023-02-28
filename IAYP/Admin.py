import openpyxl

# load the input Excel sheet
workbook = openpyxl.load_workbook('records.xlsx')
sheet = workbook.active

# create a new workbook and sheet for output
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# write headers to the output sheet
output_sheet.append(['Name', 'Wbs Number','Award Type','Email'])

# iterate over the rows in the input sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    Wbs = row[1]
    email = row[3]
    consent = row[9]
    Award_Type = row[5]

    # check if the name starts with "A" and consent is pending
    if consent == 'Pending':
        # write the data to the output sheet
        output_sheet.append([name, Wbs, Award_Type, email])

# save the output workbook
output_workbook.save('Pending_Parent_Consent.xlsx')
print("New Excell Sheet under the name Pending_Parent_Consent has been edited")
