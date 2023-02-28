import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date

# Open the Excel workbook and select the active sheet
wb = openpyxl.load_workbook('records.xlsx')
sheet = wb.active

# Get the next available row in the sheet
row = sheet.max_row + 1

# Ask the user for the details of each person
while True:
    name = input("Please enter your name - ").capitalize()
    wbs_no = input("Please enter your WBS Number - ")
    House = input("Please enter the name of your house (JA/JB/GA/GB/KA/KB/CA/CB) - ").upper()
    email = input("Please enter your email - : ").lower()
    phone_number = input("Please enter your phone number -  ")
    award_level = input("Enter award level: ").capitalize()
    dob = input("Enter date of birth in dd-mm-yyyy format: ")

    # Write the details to the Excel sheet
    sheet.cell(row=row, column=1).value = name
    sheet.cell(row=row, column=2).value = wbs_no
    sheet.cell(row=row, column=3).value = House
    sheet.cell(row=row, column=4).value = email
    sheet.cell(row=row, column=5).value = phone_number
    sheet.cell(row=row, column=6).value = award_level
    sheet.cell(row=row, column=7).value = dob
    sheet.cell(row=row, column=8).value = date.today().strftime("%d-%m-%Y")
    sheet.cell(row=row, column=9).value = "Pending"
    sheet.cell(row=row, column=10).value = "Pending"

    # Increment the row counter
    row += 1

    cont = input("Are there any other persons ! (Yes or No) - ").lower()
    if cont=="no":
        break

# Save the changes to the workbook
wb.save('records.xlsx')
