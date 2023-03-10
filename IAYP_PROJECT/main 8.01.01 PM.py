import tkinter as tk
from tkinter import messagebox
from email.mime.text import MIMEText
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import openpyxl
from datetime import date
import tkinter as tk
from PIL import Image, ImageTk

def send_parent_consent():
    def send_email(name, email, orb_number):
        print(f"{name} {email}")
        sender_email = "aadityakharkiatest@gmail.com"
        sender_password = "pxduwpgyfmtvtzmj"
        recipient_email = email

        subject = 'IAYP Parent Consent Form'
        message = f'Respected Parent,\n\nKindly find the parent consent form for {name} with ORB number {orb_number} attached.\n\nRegards,\nHariom Tripathi\n(Teacher)'

        file_path = 'Mail/Parent_Consent_IAYP-.doc'

        with open(file_path, 'rb') as f:
            file_data = f.read()
        file_name = 'Parent_Consent_IAYP.doc'

        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject

        # Add message body
        msg.attach(MIMEText(message))

        # Add attachment
        attachment = MIMEApplication(file_data, name=file_name)
        attachment['Content-Disposition'] = f'attachment; filename="{file_name}"'
        msg.attach(attachment)

        # Create SMTP session
        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()

            # Login to sender email account
            smtp.login(sender_email, sender_password)

            # Send email
            smtp.send_message(msg)

    df = pd.read_excel('Admin.xlsx')
    print(df)
    for index, row in df.iterrows():
        if row['Parent_Consent'] == 'Pending':
            name = row['Name']
            email = row['Email']
            orb_number = row['ORB_NUMBER']
            award_type = row['Award_Type']
            send_email(name, email, orb_number)
            print(f"Mail Sent to {name} at {email}")

def Fill_out_ORB_Number():
    excel_file = pd.read_excel('Admin.xlsx')
    orb_column = 'ORB_NUMBER'
    window = tk.Tk()
    window.title("ORB Number")
    def update_orb(row, orb_entry):
        orb_number = orb_entry.get()
        excel_file.loc[row, orb_column] = orb_number
        excel_file.to_excel('Admin.xlsx', index=False)
    for index, row in excel_file.iterrows():
        if pd.isna(row[orb_column]):
            # Create a label with the name and award type
            name_label = tk.Label(window, text=f"{row['Name']}, {row['Award_Type']}")
            name_label.pack(padx=50,pady=10)
            orb_entry = tk.Entry(window)
            orb_entry.pack(padx=50,pady=10)
            update_button = tk.Button(window, text="Update ORB",
                                      command=lambda row=index, orb_entry=orb_entry: update_orb(row, orb_entry))
            update_button.pack(padx=50,pady=10)
    window.mainloop()

def New_Registration():
    def submit_details():
        name = name_entry.get()
        wbs_no = wbs_no_entry.get()
        house = house_var.get()
        email = email_entry.get()
        phone = phone_entry.get()
        award_type = award_type_var.get()
        dob = dob_entry.get()

        wb = openpyxl.load_workbook('Admin.xlsx')
        sheet = wb.active
        row = sheet.max_row + 1

        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = wbs_no
        sheet.cell(row=row, column=3).value = house
        sheet.cell(row=row, column=4).value = email
        sheet.cell(row=row, column=5).value = phone
        sheet.cell(row=row, column=6).value = award_type
        sheet.cell(row=row, column=7).value = dob
        sheet.cell(row=row, column=8).value = "Pending"
        sheet.cell(row=row, column=9).value = "Pending"
        sheet.cell(row=row, column=10).value = date.today().strftime("%d-%m-%Y")
        row += 1

        wb.save('Admin.xlsx')

        # Clear the entry fields
        name_entry.delete(0, tk.END)
        wbs_no_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        phone_entry.delete(0, tk.END)
        dob_entry.delete(0, tk.END)

    window = tk.Tk()
    window.title("IAYP Registration Form")
    window.config(bg="#BAD7E9")

    # Create labels and entry fields
    name_label = tk.Label(window, text="Name:",bg="#BAD7E9")
    name_entry = tk.Entry(window,width=40)

    wbs_no_label = tk.Label(window, text="WBS NO:",bg="#BAD7E9")
    wbs_no_entry = tk.Entry(window,width=40)

    house_var = tk.StringVar(window)
    house_var.set("None")

    # Create the dropdown menu for house
    house_label = tk.Label(window, text="House:",bg="#BAD7E9")
    house_options = ["None", "Jamuna A", "Krishna A", "Ganga A", "Cauvery A", "Jamuna B", "Krishna B", "Ganga B",
                     "Cauvery B"]
    house_dropdown = tk.OptionMenu(window, house_var, *house_options)

    email_label = tk.Label(window, text="Email:",bg="#BAD7E9")
    email_entry = tk.Entry(window,width=40)

    phone_label = tk.Label(window, text="Phone Number:",bg="#BAD7E9")
    phone_entry = tk.Entry(window,width=40)

    award_type_var = tk.StringVar(window)
    award_type_var.set("Bronze")

    award_type_label = tk.Label(window, text="Award Level:",bg="#BAD7E9")
    award_type_options = ["Gold", "Silver", "Bronze"]
    award_type_dropdown = tk.OptionMenu(window, award_type_var, *award_type_options)

    dob_label = tk.Label(window, text="D.O.B.:",bg="#BAD7E9")
    dob_entry = tk.Entry(window,width=40)

    # Create the submit and quit buttons
    submit_button = tk.Button(window, text="Submit", command=submit_details,highlightthickness=0)
    quit_button = tk.Button(window, text="Quit", command=window.destroy,highlightthickness=0)

    # Arrange labels and entry fields horizontally using the grid geometry manager
    name_label.grid(row=0, column=0, sticky="E")
    name_entry.grid(row=0, column=1, padx=10, pady=10)

    wbs_no_label.grid(row=1, column=0, sticky="E")
    wbs_no_entry.grid(row=1, column=1, padx=10, pady=10)

    house_label.grid(row=2, column=0, sticky="E")
    house_dropdown.grid(row=2, column=1, padx=10, pady=10)

    email_label.grid(row=3, column=0, sticky="E")
    email_entry.grid(row=3, column=1, padx=10, pady=10)

    phone_label.grid(row=4, column=0, sticky="E")
    phone_entry.grid(row=4, column=1, padx=10, pady=10)

    award_type_label.grid(row=5, column=0, sticky="E")
    award_type_dropdown.grid(row=5, column=1, padx=10, pady=10)

    dob_label.grid(row=6, column=0, sticky="E")
    dob_entry.grid(row=6, column=1, padx=10, pady=10)

    submit_button.grid(row=7, column=1)
    quit_button.grid(row=7, column=2, padx=10, pady=10)

    # Run the GUI
    window.mainloop()

def get_pending_payment():

    def update_payment_status():
        excel_file = pd.read_excel('Admin.xlsx')
        for index, row in excel_file.iterrows():
            if row['Payment_Status'] == 'Pending':
                excel_file.at[index, 'Payment_Status'] = 'Completed'

        # Save the updated Excel file
        excel_file.to_excel('Admin.xlsx', index=False)

        # Show a message box confirming the update
        tk.messagebox.showinfo("Payment Status Updated", "All pending payments have been marked as completed.")

    # Function to display the payment status
    def display_payment_status():
        # Read the data from the Excel file
        excel_file = pd.read_excel('Admin.xlsx')

        # Create dictionaries to store data for each table
        bronze_dict = {'Name': [], 'WBS_NO': [], 'House': []}
        silver_dict = {'Name': [], 'WBS_NO': [], 'House': []}
        gold_dict = {'Name': [], 'WBS_NO': [], 'House': []}

        # Loop through the data and insert into the appropriate dictionary
        for index, row in excel_file.iterrows():
            if row['Payment_Status'] == 'Pending':
                if row['Award_Type'] == 'Bronze':
                    bronze_dict['Name'].append(row['Name'])
                    bronze_dict['WBS_NO'].append(row['WBS_NO'])
                    bronze_dict['House'].append(row['House'])
                elif row['Award_Type'] == 'Silver':
                    silver_dict['Name'].append(row['Name'])
                    silver_dict['WBS_NO'].append(row['WBS_NO'])
                    silver_dict['House'].append(row['House'])
                elif row['Award_Type'] == 'Gold':
                    gold_dict['Name'].append(row['Name'])
                    gold_dict['WBS_NO'].append(row['WBS_NO'])
                    gold_dict['House'].append(row['House'])

        # Create a Tkinter window
        window = tk.Tk()
        window.title("Payment Status")

        # Create a function to display the table and count for each category
        def display_table(category, data):
            # Create a frame for the table
            table_frame = tk.Frame(window)
            table_frame.pack(pady=10)

            # Create a label for the category and count
            tk.Label(table_frame, text=f"{category} ({len(data['Name'])})").pack()

            # Create a table using a Text widget
            table = tk.Text(table_frame, height=10, width=50)
            table.pack()

            # Insert data into the table
            for i in range(len(data['Name'])):
                table.insert(tk.END, f"{data['Name'][i]}\t{data['WBS_NO'][i]}\t{data['House'][i]}\n")

        # Display the tables
        display_table("Bronze", bronze_dict)
        display_table("Silver", silver_dict)
        display_table("Gold", gold_dict)

        # Add a button to update the payment status
        update_button = tk.Button(window, text="Completed", command=update_payment_status)
        update_button.pack(pady=10)

        # Start the Tkinter event loop
        window.mainloop()

    # Display the payment status
    display_payment_status()

def all_data():
    excel_file = pd.read_excel('Admin.xlsx')

    # Create dictionaries to store data for each table
    bronze_dict = {'Name': [], 'WBS_NO': [], 'House': [],'Email':[],'Phone Number':[],'ORB_NUMBER':[],"Registration_Date":[]}
    silver_dict = {'Name': [], 'WBS_NO': [], 'House': [],'Email':[],'Phone Number':[],'ORB_NUMBER':[],"Registration_Date":[]}
    gold_dict = {'Name': [], 'WBS_NO': [], 'House': [],'Email':[],'Phone Number':[],'ORB_NUMBER':[],"Registration_Date":[]}

    # Loop through the data and insert into the appropriate dictionary
    for index, row in excel_file.iterrows():
        if row['Award_Type'] == 'Bronze':
            bronze_dict['Name'].append(row['Name'])
            bronze_dict['WBS_NO'].append(row['WBS_NO'])
            bronze_dict['House'].append(row['House'])
            bronze_dict['Email'].append(row['Email'])
            bronze_dict['Phone Number'].append(row['Phone Number'])
            bronze_dict['ORB_NUMBER'].append(row['ORB_NUMBER'])
            bronze_dict['Registration_Date'].append(row['Registration_Date'])

        elif row['Award_Type'] == 'Silver':
            silver_dict['Name'].append(row['Name'])
            silver_dict['WBS_NO'].append(row['WBS_NO'])
            silver_dict['House'].append(row['House'])
            silver_dict['Email'].append(row['Email'])
            silver_dict['Phone Number'].append(row['Phone Number'])
            silver_dict['ORB_NUMBER'].append(row['ORB_NUMBER'])
            silver_dict['Registration_Date'].append(row['Registration_Date'])

        elif row['Award_Type'] == 'Gold':
            gold_dict['Name'].append(row['Name'])
            gold_dict['WBS_NO'].append(row['WBS_NO'])
            gold_dict['House'].append(row['House'])
            gold_dict['Email'].append(row['Email'])
            gold_dict['Phone Number'].append(row['Phone Number'])
            gold_dict['ORB_NUMBER'].append(row['ORB_NUMBER'])
            silver_dict['Registration_Date'].append(row['Registration_Date'])

    window = tk.Tk()
    window.title("All Data")

    def display_table(category, data):
        # Create a frame for the table
        table_frame = tk.Frame(window)
        table_frame.pack(pady=10)

        # Create a label for the category and count
        tk.Label(table_frame, text=f"{category} ({len(data['Name'])})").pack()

        # Create a table using a Text widget
        table = tk.Text(table_frame, height=10, width=150)
        table.pack()

        table.insert(tk.END,
                     "-------------------------------------------------------------------------------------------------------------------------------------------\n")
        table.insert(tk.END, "Name\t\t\tWBS NO.\t\tHouse\t\tEmail\t\t\t\tPhone Number\t\tORB NUMBER\t\tRegistration Date\n\n")
        table.insert(tk.END,
                     "--------------------------------------------------------------------------------------------------------------------------------------------\n")

        # Insert data into the table
        for i in range(len(data['Name'])):
            table.insert(tk.END, f"{data['Name'][i]}\t\t\t{data['WBS_NO'][i]}\t\t{data['House'][i]}\t\t{data['Email'][i]}\t\t\t\t{data['Phone Number'][i]}\t\t{data['ORB_NUMBER'][i]}\t\t{data['Registration_Date'][i]}\n\n")

    # Display the tables
    display_table("Bronze", bronze_dict)
    display_table("Silver", silver_dict)
    display_table("Gold", gold_dict)

    # Start the Tkinter event loop
    window.mainloop()

def update_pending_consent():
    import pandas as pd
    import tkinter as tk

    # Load the Excel file using pandas
    df = pd.read_excel('Admin.xlsx', sheet_name='Sheet1')

    # Create a tkinter window
    root = tk.Tk()
    root.title("Parent Consent Updation")
    root.minsize(400, 400)

    # Create a frame to hold the labels and buttons
    frame = tk.Frame(root)
    frame.pack()

    # Loop through all rows in the dataframe
    for index, row in df.iterrows():
        if row['Parent_Consent'] == 'Pending':
            # Extract name and award type
            name = row['Name']
            award_type = row['Award_Type']

            # Create a label to display the name and award type
            label = tk.Label(frame, text=f'{name} - {award_type}', padx=20, pady=20)
            label.grid(row=index, column=0)

            # Create a button to mark the row as complete
            def mark_complete(index=index):
                df.loc[index, 'Parent_Consent'] = 'Complete'
                df.to_excel('Admin.xlsx', index=False)

            button = tk.Button(frame, text='Complete', command=mark_complete)
            button.grid(row=index, column=1)

    # Run the tkinter event loop
    root.mainloop()

root = tk.Tk()
root.title("IAYP Application")
root.maxsize(700,600)
root.minsize(700,600)
root.configure(background="#ADD8E6")

# create and place logos on top of the window
logo1 = tk.PhotoImage(file="iayp_logo.png").subsample(1, 1)
label1 = tk.Label(root, image=logo1,background="#ADD8E6")
label1.grid(row=0, column=0, padx=20, pady=10)

logo2 = tk.PhotoImage(file="welham.png").subsample(2, 2)
label2 = tk.Label(root, image=logo2)
label2.grid(row=0, column=1, padx=20, pady=10)

frame1 = tk.Frame(root, width=200, height=200,background="#ADD8E6",)
frame1.grid(row=1, column=0)

frame2 = tk.Frame(root, width=200, height=200,background="#ADD8E6")
frame2.grid(row=1, column=1)

frame3 = tk.Frame(root, width=200, height=200,background="#ADD8E6")
frame3.grid(row=2, column=0)

frame4 = tk.Frame(root, width=200, height=200,background="#ADD8E6")
frame4.grid(row=2, column=1)

frame5 = tk.Frame(root, width=200, height=200,background="#ADD8E6")
frame5.grid(row=3, column=0)

frame6 = tk.Frame(root, width=200, height=200,background="#ADD8E6")
frame6.grid(row=3, column=1)

# create and place buttons with enlarged text in each quadrant
button_send_parent_consent = tk.Button(frame1, text="Send Parent Consent", command=send_parent_consent,font=("Arial", 20))
button_send_parent_consent.pack(pady=5e0,padx=80)

button_fill_out_ORB_number = tk.Button(frame2, text="Fill out ORB Number", command=Fill_out_ORB_Number,font=("Arial", 20))
button_fill_out_ORB_number.pack(pady=50,padx=80)

button_new_registration = tk.Button(frame3, text="New Registration", command=New_Registration,font=("Arial", 20))
button_new_registration.pack(pady=50,padx=80)

button_get_pending_payment = tk.Button(frame4, text="Get Pending Payment", command=get_pending_payment,font=("Arial", 20))
button_get_pending_payment.pack(pady=50,padx=80)

button_send_all_data = tk.Button(frame5, text="All Data", command=all_data,font=("Arial", 20))
button_send_all_data.pack(pady=50,padx=80)

button_complete_parent_consent = tk.Button(frame6, text="Parent Consent Updation", command=update_pending_consent,font=("Arial", 20))
button_complete_parent_consent.pack(pady=50,padx=80)

root.mainloop()
